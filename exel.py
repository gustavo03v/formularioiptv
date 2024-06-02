import openpyxl
from openpyxl import Workbook

# Função para salvar dados no Excel
def save_to_excel(data):
    try:
        # Tente carregar o workbook existente
        workbook = openpyxl.load_workbook('assinaturas.xlsx')
    except FileNotFoundError:
        # Se o arquivo não existir, crie um novo workbook
        workbook = Workbook()
    
    sheet = workbook.active
    sheet.title = "Assinaturas"

    # Verifique se o cabeçalho já existe
    if sheet.max_row == 1 and sheet.max_column == 1:
        sheet.append(["Nome", "Email", "Plano", "Pagamento"])

    # Adicione os dados
    sheet.append([data['nome'], data['email'], data['plano'], "Pix: be029f35-ed1c-493c-9278-52e9bca7f10e"])

    # Salve o workbook
    workbook.save('assinaturas.xlsx')

# Exemplo de dados recebidos do formulário
form_data = {
    "nome": "João Silva",
    "email": "joao.silva@email.com",
    "plano": "trimestral"
}

# Salve os dados no Excel
save_to_excel(form_data)
